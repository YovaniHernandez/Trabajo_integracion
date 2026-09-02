from dataclasses import dataclass
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from config.categories.registry import registry

@dataclass
class ProductRow:
    sheet_name: str
    row_idx: int
    sku: str
    title: str
    linea: str
    familia: str
    subfamilia: str
    unidad: str
    image_url: str
    raw_values: Dict[str, Any]
    is_completed: bool = False

class ExcelReader:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
        self.wb = openpyxl.load_workbook(self.file_path, data_only=True)

    @property
    def sheet_names(self) -> List[str]:
        return self.wb.sheetnames

    def get_column_mapping(self, ws: Worksheet, header_row: int = 1) -> Dict[str, int]:
        """Maps column name/header to 1-indexed column index."""
        col_map = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col_idx).value
            if val is not None:
                header_str = str(val).strip()
                col_map[header_str] = col_idx
        return col_map

    def read_sheet(self, sheet_name: str, force: bool = False) -> List[ProductRow]:
        """Reads product rows starting at row 5 (skipping row 1-4 metadata)."""
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {self.file_path}")

        ws = self.wb[sheet_name]
        col_map = self.get_column_mapping(ws)
        category_cfg = registry.get_by_sheet(sheet_name)
        start_row = category_cfg.get("start_row", 5) if category_cfg else 5

        sku_col = col_map.get("CODIGO", 1)
        title_col = col_map.get("DESCRIPCION", 2)
        desc_col = col_map.get("Descripción", col_map.get("DESCRIPCION_COMERCIAL", None))

        rows: List[ProductRow] = []

        for r in range(start_row, ws.max_row + 1):
            sku_val = ws.cell(r, sku_col).value
            title_val = ws.cell(r, title_col).value

            # If both SKU and Title are empty, skip
            if sku_val is None and title_val is None:
                continue

            sku = str(sku_val).strip() if sku_val is not None else ""
            title = str(title_val).strip() if title_val is not None else ""

            if not sku and not title:
                continue

            # Collect raw values for all known columns
            raw_values: Dict[str, Any] = {}
            for col_name, c_idx in col_map.items():
                val = ws.cell(r, c_idx).value
                raw_values[col_name] = val

            # Check if row is already fully completed
            has_description = False
            if desc_col and ws.cell(r, desc_col).value:
                desc_text = str(ws.cell(r, desc_col).value).strip()
                if len(desc_text) > 20 and not desc_text.lower().startswith("descripción optimizada"):
                    has_description = True

            marca_val = str(raw_values.get("Marca", "") or "").strip()
            is_generic_brand = (marca_val.lower() == "genérico" or marca_val.lower() == "generico")

            # Check category specific required fields
            cat_lower = sheet_name.lower()
            specs_complete = True
            if "salpicadera" in cat_lower:
                alt = raw_values.get("Altura")
                esp = raw_values.get("Espesor")
                mat = raw_values.get("Material")
                if not alt or not esp or not mat or str(alt).lower() == "obligatorio" or str(esp).lower() == "obligatorio":
                    specs_complete = False
            elif "sprocket" in cat_lower:
                largo = raw_values.get("Largo de la cadena")
                mat_sprocket = str(raw_values.get("Material del sprocket", "") or "").strip()
                if not largo or mat_sprocket.lower() != "acero" or str(largo).lower() == "obligatorio":
                    specs_complete = False
            elif "barra" in cat_lower:
                diam = raw_values.get("Diámetro")
                largo = raw_values.get("Largo")
                if not diam or not largo or str(diam).lower() == "obligatorio" or str(largo).lower() == "obligatorio":
                    specs_complete = False

            is_completed = (not force) and has_description and is_generic_brand and specs_complete

            p_row = ProductRow(
                sheet_name=sheet_name,
                row_idx=r,
                sku=sku,
                title=title,
                linea=str(raw_values.get("LINEA", "") or ""),
                familia=str(raw_values.get("FAMILIA", "") or ""),
                subfamilia=str(raw_values.get("SUB-FAMILIA", "") or ""),
                unidad=str(raw_values.get("UNIDAD", "") or ""),
                image_url=str(raw_values.get("URL Imagen", "") or ""),
                raw_values=raw_values,
                is_completed=is_completed,
            )
            rows.append(p_row)

        return rows
