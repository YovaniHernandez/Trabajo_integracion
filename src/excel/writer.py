import shutil
from pathlib import Path
from typing import Dict, Any, Optional
import openpyxl

class ExcelWriter:
    def __init__(self, source_path: str | Path, output_path: Optional[str | Path] = None):
        self.source_path = Path(source_path)
        self.output_path = Path(output_path) if output_path else self.source_path

        # If output path is different and doesn't exist yet, copy source file
        if self.output_path != self.source_path and not self.output_path.exists():
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(self.source_path, self.output_path)

        # Load editable workbook
        target_file = self.output_path if self.output_path.exists() else self.source_path
        self.wb = openpyxl.load_workbook(target_file)
        self._col_maps: Dict[str, Dict[str, int]] = {}

    def _get_col_map(self, sheet_name: str) -> Dict[str, int]:
        if sheet_name in self._col_maps:
            return self._col_maps[sheet_name]

        ws = self.wb[sheet_name]
        col_map: Dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(1, col_idx).value
            if val is not None and str(val).strip():
                col_map[str(val).strip()] = col_idx

        self._col_maps[sheet_name] = col_map
        return col_map

    def update_row(self, sheet_name: str, row_idx: int, updates: Dict[str, Any], save_now: bool = True):
        """Updates specific columns in a sheet at row_idx and optionally saves the workbook."""
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

        ws = self.wb[sheet_name]
        col_map = self._get_col_map(sheet_name)

        for col_name, value in updates.items():
            if col_name in col_map:
                col_idx = col_map[col_name]
                ws.cell(row=row_idx, column=col_idx, value=value)

        if save_now:
            self.save()

    def save(self):
        """Persists changes to the output Excel file."""
        self.wb.save(self.output_path)
