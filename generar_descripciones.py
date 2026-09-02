"""
Generador de descripciones para MercadoLibre (Iron Racing)
============================================================

Lee un Excel con SKU (columna A) y Título del artículo (columna B),
manda cada título a la API de Claude (con la herramienta de búsqueda
web activada para complementar datos técnicos del producto), y escribe
la descripción generada en la columna C del mismo archivo.

REQUISITOS
----------
1. Python 3.9+
2. Instalar dependencias:
       pip install anthropic openpyxl
3. Tener una API key de Anthropic (https://console.anthropic.com/settings/keys)
   y exportarla como variable de entorno:

       # Mac/Linux
       export ANTHROPIC_API_KEY="tu-api-key"

       # Windows (PowerShell)
       $env:ANTHROPIC_API_KEY="tu-api-key"

USO
---
    python generar_descripciones.py entrada.xlsx

Opcional:
    python generar_descripciones.py entrada.xlsx --hoja "Hoja1" --salida salida.xlsx

El script:
- Lee cada fila desde la fila 2 (asume fila 1 = encabezados).
- Si la columna C ya tiene texto, se salta esa fila (para poder correr
  el script varias veces sin repetir trabajo ni gastar créditos).
- Guarda el archivo después de cada fila, así que si se interrumpe no
  pierdes el avance.
"""

import argparse
import sys
import time

from openpyxl import load_workbook
import anthropic

# ---------------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------------

MODEL = "claude-sonnet-5"          # Buena relación calidad/costo. Alternativas:
                                    #   "claude-opus-4-8"  -> más calidad, más caro
                                    #   "claude-haiku-4-5-20251001" -> más barato/rápido
MAX_TOKENS = 1500
COLUMNA_SKU = 1      # A
COLUMNA_TITULO = 2   # B
COLUMNA_DESCRIPCION = 3  # C
FILA_INICIO = 2       # asume que la fila 1 son encabezados
PAUSA_ENTRE_LLAMADAS = 1.0  # segundos, para no saturar la API

# Ejemplos de estilo que sirven como referencia para el modelo.
# (Los mismos que compartiste, se usan tal cual como "few-shot examples".)
EJEMPLOS_ESTILO = """
Casco: CASCO MOTOCROSS IRON RACING – LÍNEA APOLO ¡Diseño agresivo, seguridad garantizada y estilo único! Prepárate para dominar cualquier terreno con la línea de cascos Apolo de Iron Racing, ideales para motocross, enduro o uso urbano. Su estructura robusta y diseño aerodinámico brindan una protección superior sin sacrificar comodidad ni estilo.

** Características destacadas:
-Certificación DOT (cumple con estándares internacionales de seguridad)
-Fabricado en policarbonato de alta resistencia
-Interior acolchado, desmontable y lavable
-Ventilación frontal, superior y trasera para máxima frescura
-Visor alargado tipo cross con diseño deportivo
-Sistema de cierre con hebilla micrométrica
-Tallas: CH, M, G, EG (según disponibilidad)
-Disponible en colores vibrantes y llamativos

Modelos con combinaciones en: Azul Neón, Amarillo Neón, Violeta, y Rojo.
Ideal para motociclistas que buscan impactar visualmente mientras se mantienen protegidos en cada ruta.

GUÍA DE TALLAS: CH: 55-56 cm / M: 57-58 cm / L: 59-60 cm / XL: 60-61 cm

---

Caja para motos: Caja Porta-Equipaje IRON RACING – 55 Litros
Lleva todo lo que necesitas con seguridad y estilo con las cajas porta-equipaje Iron Racing, diseñadas especialmente para motociclistas exigentes. Fabricadas en plástico resistente y con diseño cuadrado, estas cajas ofrecen gran capacidad, durabilidad y un toque moderno para tu moto.

Características generales:
-Material: Plástico rígido de alta resistencia
-Dimensiones: 44 × 34.5 × 37.5 cm
-Capacidad: 55 litros
-Diseño cuadrado con refuerzos en esquinas
-Sistema de cierre metálico con llave para mayor seguridad

Colores disponibles:
*Arena – Modelo CAJ-7104-0037
*Verde Oscuro Militar – Modelo CAJ-7104-0038
*Azul Oscuro – Modelo CAJ-7104-0039

Ideales para viajes largos, mensajería o uso diario. ¡Complementa tu motocicleta con el estilo y funcionalidad de Iron Racing!

---

Guantes: Guantes Motociclismo Táctil Touch Iron Racing Phoenix Azul
Iron Racing, marca líder en accesorios para motocicletas, ofrece productos con altos estándares de calidad, aprobados en laboratorios internacionales para asegurar su efectividad.

Estos guantes no solo son atractivos, sino que también brindan seguridad en tus viajes cortos y largos en tu máquina veloz.

Características:
- Diseño táctil para uso de dispositivos móviles
- Material resistente y duradero
- Protección en nudillos y palma
- Ajuste cómodo y seguro
- Ideal para motociclismo y actividades al aire libre

¡Prepárate para disfrutar de la experiencia Iron Racing!
""".strip()

SYSTEM_PROMPT = f"""Eres redactor de fichas de producto para publicaciones de MercadoLibre de la marca IRON RACING (accesorios y equipo para motociclistas: cascos, guantes, cajas porta-equipaje, etc.).

Debes escribir descripciones con el MISMO estilo, tono y estructura que estos ejemplos reales de la marca:

{EJEMPLOS_ESTILO}

REGLAS DE ESTILO A SEGUIR SIEMPRE:
1. Título llamativo en mayúsculas al inicio (nombre del producto + línea/modelo si aplica).
2. Un párrafo de introducción tipo "gancho de venta" (agresivo/emocional, resalta seguridad + estilo).
3. Sección "Características destacadas" o "Características:" con viñetas (guiones), datos técnicos concretos: materiales, certificaciones, medidas, capacidad, ventilación, cierres, etc.
4. Si aplica, sección de colores/modelos disponibles.
5. Si aplica, guía de tallas o medidas.
6. Cierre motivador y de marca (ej. "¡Prepárate para disfrutar de la experiencia Iron Racing!", "Estilo, Seguridad y Confort.").
7. Usa SIEMPRE la herramienta de búsqueda web para investigar el producto real (a partir del título/SKU) y así completar datos técnicos plausibles y verídicos: materiales típicos, certificaciones (DOT, ECE, etc.), tallas estándar, dimensiones, capacidades, tipos de cierre, etc. Si no encuentras el producto exacto, básate en productos equivalentes de la misma categoría para dar información técnica realista, sin inventar certificaciones o datos que no puedas sustentar.
8. No inventes números de modelo o certificaciones que no existan; si no hay información confiable, usa términos generales ("material resistente", "diseño ergonómico") en lugar de cifras falsas.
9. Escribe en español, tono comercial mexicano, igual que los ejemplos.
10. Devuelve ÚNICAMENTE el texto final de la descripción, sin comentarios adicionales, sin explicaciones de tu proceso, sin markdown de encabezados (#), listo para copiar y pegar en MercadoLibre.
11. MUY IMPORTANTE: tu respuesta de texto debe EMPEZAR directamente con el título del producto en mayúsculas. Nunca escribas frases de transición antes del título como "Ahora tengo suficiente información...", "Con esta información puedo redactar...", "Aquí está la descripción:", etc. La primera línea de tu respuesta es siempre el título del producto, sin nada antes.
"""


def limpiar_descripcion(texto: str) -> str:
    """
    Respaldo por si el modelo agrega alguna frase de transición antes del
    título (ej. "Ahora tengo suficiente información para redactar la ficha.").
    Busca la primera línea con "cara de título" (mayormente en mayúsculas,
    con longitud razonable) y descarta todo lo anterior.
    """
    lineas = texto.strip().split("\n")

    for i, linea in enumerate(lineas):
        candidata = linea.strip()
        letras = [c for c in candidata if c.isalpha()]
        if not letras:
            continue
        proporcion_mayusculas = sum(1 for c in letras if c.isupper()) / len(letras)
        # Una línea "título" típica: en mayúsculas casi en su totalidad y con
        # cierta longitud (para no confundir con siglas sueltas).
        if proporcion_mayusculas > 0.8 and len(candidata) >= 15:
            return "\n".join(lineas[i:]).strip()

    # Si no se encontró una línea con cara de título, se devuelve tal cual.
    return texto.strip()


def generar_descripcion(client: anthropic.Anthropic, sku: str, titulo: str) -> str:
    """Llama a la API de Claude (con web_search) para generar una descripción."""
    user_prompt = (
        f"Genera la descripción de MercadoLibre para este artículo de Iron Racing.\n\n"
        f"SKU: {sku}\n"
        f"Título / nombre del artículo: {titulo}\n\n"
        f"Investiga en línea si es necesario para obtener datos técnicos reales de este "
        f"producto o de productos equivalentes de esa categoría, y luego escribe la "
        f"descripción siguiendo el estilo indicado."
    )

    response = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        system=SYSTEM_PROMPT,
        tools=[{"type": "web_search_20250305", "name": "web_search"}],
        messages=[{"role": "user", "content": user_prompt}],
    )

    # La respuesta puede incluir bloques de tipo "text" y bloques de uso de
    # herramienta (búsquedas). Solo nos interesa concatenar el texto final.
    partes_texto = [bloque.text for bloque in response.content if bloque.type == "text"]
    texto_completo = "\n".join(partes_texto).strip()
    return limpiar_descripcion(texto_completo)


def main():
    parser = argparse.ArgumentParser(description="Genera descripciones de MercadoLibre con Claude.")
    parser.add_argument("archivo", help="Ruta al archivo .xlsx de entrada (columna A = SKU, columna B = título)")
    parser.add_argument("--hoja", default=None, help="Nombre de la hoja a usar (por defecto, la hoja activa)")
    parser.add_argument("--salida", default=None, help="Ruta de salida (por defecto sobrescribe el mismo archivo)")
    args = parser.parse_args()

    salida = args.salida or args.archivo

    print(f"Abriendo {args.archivo} ...")
    wb = load_workbook(args.archivo)
    ws = wb[args.hoja] if args.hoja else wb.active

    client = anthropic.Anthropic()  # usa ANTHROPIC_API_KEY del entorno

    fila = FILA_INICIO
    procesadas = 0
    saltadas = 0

    while True:
        sku_cell = ws.cell(row=fila, column=COLUMNA_SKU)
        titulo_cell = ws.cell(row=fila, column=COLUMNA_TITULO)

        # Detener cuando ya no hay más filas con datos
        if sku_cell.value is None and titulo_cell.value is None:
            break

        sku = str(sku_cell.value).strip() if sku_cell.value is not None else ""
        titulo = str(titulo_cell.value).strip() if titulo_cell.value is not None else ""

        desc_cell = ws.cell(row=fila, column=COLUMNA_DESCRIPCION)

        if not titulo:
            fila += 1
            continue

        if desc_cell.value:
            print(f"Fila {fila} ({sku}): ya tiene descripción, se salta.")
            saltadas += 1
            fila += 1
            continue

        print(f"Fila {fila} ({sku}): generando descripción para '{titulo}' ...")
        try:
            descripcion = generar_descripcion(client, sku, titulo)
            desc_cell.value = descripcion
            wb.save(salida)  # guarda progreso después de cada fila
            procesadas += 1
            print(f"  -> OK ({len(descripcion)} caracteres)")
        except Exception as e:
            print(f"  -> ERROR en fila {fila}: {e}")
            print("  Se continúa con la siguiente fila; puedes volver a correr el script después para reintentar esta.")

        fila += 1
        time.sleep(PAUSA_ENTRE_LLAMADAS)

    print(f"\nListo. Filas procesadas: {procesadas}. Filas saltadas (ya tenían descripción): {saltadas}.")
    print(f"Archivo guardado en: {salida}")


if __name__ == "__main__":
    sys.exit(main())
